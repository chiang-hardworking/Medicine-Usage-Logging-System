import customtkinter as ctk
import tkinter.ttk as ttk
import tkinter.messagebox as messagebox
import serial
import serial.tools.list_ports
import threading
import datetime
import os
import json
import re
import sys
import webbrowser
from urllib import error as url_error
from urllib import request as url_request
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

CONFIG_FILE = "config.json"
VERSION_FILE = "version.json"


class BalanceGUI(ctk.CTk):
    APP_NAME = "Medicine Usage Logging System"
    REPO_OWNER = "chiang-hardworking"
    REPO_NAME = "Medicine-Usage-Logging-System"
    RELEASES_LATEST_API = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/releases/latest"
    RELEASES_PAGE = f"https://github.com/{REPO_OWNER}/{REPO_NAME}/releases/latest"

    DETAIL_SHEET = "詳細記錄"
    INVENTORY_SHEET = "庫存餘量"
    LEGACY_DETAIL_SHEET = "詳細記錄(舊版)"
    LEGACY_INVENTORY_SHEET = "庫存餘量(舊版)"
    NO_PORT_TEXT = "找不到可用序列埠"
    TEST_MODE_PORT = "測試模式 (模擬資料)"
    TEST_UNLOCK_SHORTCUT = "<Control-Shift-T>"
    IN_TYPE = "入庫"
    OUT_TYPE = "出庫"

    DETAIL_HEADERS = ["員工編號", "日期", "時間", "產品編號", "名稱", "類別", "重量", "瓶", "餘量"]
    INVENTORY_HEADERS = ["ID", "名稱", "單位", "標準", "餘量"]

    def __init__(self):
        super().__init__()

        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
        self.config = self.load_config()
        self.current_version = self.load_local_version()
        self.title(f"METTLER TOLEDO 電子秤資料處理系統 v{self.current_version}")
        self.geometry("980x800")
        self.file_name = "medicine_data.xlsx"
        self.is_monitoring = False
        self.is_test_mode = False
        self.test_mode_unlocked = False
        self.ser = None

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar = ctk.CTkFrame(self, width=250, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")

        self.logo_label = ctk.CTkLabel(
            self.sidebar, text="秤重資料管理", font=ctk.CTkFont(size=20, weight="bold")
        )
        self.logo_label.pack(pady=15, padx=10)

        self.port_label = ctk.CTkLabel(self.sidebar, text="COM Port:")
        self.port_label.pack(pady=(5, 0))
        available_ports = self.get_ports()
        self.port_menu = ctk.CTkOptionMenu(self.sidebar, values=available_ports)
        self.port_menu.pack(pady=5, padx=20)

        last_port = self.config.get("last_port")
        if last_port in available_ports:
            self.port_menu.set(last_port)

        self.emp_label = ctk.CTkLabel(self.sidebar, text="員工編號:")
        self.emp_label.pack(pady=(5, 0))
        self.emp_entry = ctk.CTkEntry(self.sidebar)
        self.emp_entry.pack(pady=5, padx=20)
        self.emp_entry.insert(0, self.config.get("last_employee_id", ""))

        self.id_label = ctk.CTkLabel(self.sidebar, text="產品編號:")
        self.id_label.pack(pady=(5, 0))
        self.id_entry = ctk.CTkEntry(self.sidebar)
        self.id_entry.pack(pady=5, padx=20)
        self.id_entry.insert(0, "0000")

        self.prod_label = ctk.CTkLabel(self.sidebar, text="名稱:")
        self.prod_label.pack(pady=(5, 0))
        self.prod_combo = ctk.CTkComboBox(self.sidebar, values=self.config.get("product_history", []))
        self.prod_combo.pack(pady=5, padx=20)

        self.bot_label = ctk.CTkLabel(self.sidebar, text="瓶數:")
        self.bot_label.pack(pady=(5, 0))
        self.bot_entry = ctk.CTkEntry(self.sidebar)
        self.bot_entry.pack(pady=5, padx=20)
        self.bot_entry.insert(0, "1")

        self.type_label = ctk.CTkLabel(self.sidebar, text="類別:")
        self.type_label.pack(pady=(15, 0))
        self.type_var = ctk.StringVar(value=self.OUT_TYPE)
        self.type_seg = ctk.CTkSegmentedButton(
            self.sidebar, values=[self.IN_TYPE, self.OUT_TYPE], variable=self.type_var
        )
        self.type_seg.pack(pady=5, padx=20)

        self.btn_mod_std = ctk.CTkButton(
            self.sidebar,
            text="修改當前產品編號標準",
            command=self.modify_standard,
            fg_color="#b58d35",
            hover_color="#8a6a25",
        )
        self.btn_mod_std.pack(pady=(20, 0), padx=20)

        self.btn_start = ctk.CTkButton(
            self.sidebar,
            text="開始監聽",
            command=self.toggle_monitoring,
            fg_color="green",
            hover_color="#006400",
        )
        self.btn_start.pack(pady=(30, 20), padx=20)

        self.version_label = ctk.CTkLabel(self.sidebar, text=f"版本: {self.current_version}")
        self.version_label.pack(pady=(0, 6))
        self.btn_check_update = ctk.CTkButton(
            self.sidebar,
            text="檢查更新",
            command=self.check_for_updates,
            fg_color="#1f538d",
            hover_color="#14375e",
        )
        self.btn_check_update.pack(pady=(0, 14), padx=20)

        self.test_frame = ctk.CTkFrame(self.sidebar, fg_color="#2f2f2f")
        self.test_weight_label = ctk.CTkLabel(self.test_frame, text="測試重量:")
        self.test_weight_entry = ctk.CTkEntry(self.test_frame)
        self.test_unit_label = ctk.CTkLabel(self.test_frame, text="單位:")
        self.test_unit_entry = ctk.CTkEntry(self.test_frame)
        self.test_send_btn = ctk.CTkButton(
            self.test_frame, text="送出測試數據", command=self.submit_test_data, fg_color="#3d8c40", hover_color="#2f6b31"
        )

        self.main_content = ctk.CTkFrame(self)
        self.main_content.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        self.main_content.grid_rowconfigure(2, weight=1)
        self.main_content.grid_columnconfigure(0, weight=1)

        self.weight_frame = ctk.CTkFrame(self.main_content, fg_color="transparent")
        self.weight_frame.grid(row=0, column=0, pady=(10, 0))
        self.weight_display = ctk.CTkLabel(
            self.weight_frame, text="0.0000", font=ctk.CTkFont(size=70, weight="bold")
        )
        self.weight_display.pack()
        self.unit_label = ctk.CTkLabel(self.weight_frame, text="---", font=ctk.CTkFont(size=25))
        self.unit_label.pack(pady=(0, 10))

        self.log_box = ctk.CTkTextbox(self.main_content, height=100)
        self.log_box.grid(row=1, column=0, pady=(0, 10), padx=20, sticky="ew")

        self.setup_treeview()

        # 啟動時自動升級舊版工作表格式（只調整「詳細記錄」）
        self.upgrade_workbook_structure()
        self.load_recent_records()
        self.bind_all(self.TEST_UNLOCK_SHORTCUT, self.handle_test_mode_shortcut)

    def setup_treeview(self):
        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Treeview",
            background="#2b2b2b",
            foreground="white",
            fieldbackground="#2b2b2b",
            borderwidth=0,
            rowheight=25,
        )
        style.configure(
            "Treeview.Heading",
            background="#3b3b3b",
            foreground="white",
            relief="flat",
            font=("Arial", 10, "bold"),
        )
        style.map("Treeview", background=[("selected", "#1f538d")])

        self.tree_frame = ctk.CTkFrame(self.main_content)
        self.tree_frame.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.tree_frame.grid_columnconfigure(0, weight=1)
        self.tree_frame.grid_rowconfigure(0, weight=1)

        columns = tuple(self.DETAIL_HEADERS)
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", height=10)

        widths = [100, 95, 85, 95, 120, 75, 90, 50, 90]
        for col, width in zip(columns, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")

        self.tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=0, column=1, sticky="ns")

    def load_recent_records(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if not os.path.exists(self.file_name):
            return

        try:
            wb = load_workbook(self.file_name, read_only=True)
            ws = self.find_detail_sheet(wb)
            if ws:
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    recent_data = rows[1:][-10:]
                    for row in reversed(recent_data):
                        normalized = self.normalize_detail_row(row)
                        if normalized:
                            self.tree.insert("", "end", values=normalized)
            wb.close()
        except Exception as exc:
            self.update_log(f"讀取最近紀錄失敗: {exc}")

    def load_local_version(self):
        version_path = os.path.join(self.base_dir, VERSION_FILE)
        fallback_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), VERSION_FILE)
        for path in [version_path, fallback_path]:
            if not os.path.exists(path):
                continue
            try:
                with open(path, "r", encoding="utf-8") as file:
                    data = json.load(file)
                    return str(data.get("version", "0.0.0"))
            except Exception:
                continue
        return "0.0.0"

    def normalize_version(self, version_text):
        cleaned = str(version_text).strip().lstrip("vV")
        parts = re.findall(r"\d+", cleaned)
        if not parts:
            return (0, 0, 0)
        numbers = [int(x) for x in parts[:3]]
        while len(numbers) < 3:
            numbers.append(0)
        return tuple(numbers)

    def fetch_latest_release(self):
        req = url_request.Request(
            self.RELEASES_LATEST_API,
            headers={"Accept": "application/vnd.github+json", "User-Agent": self.APP_NAME},
        )
        with url_request.urlopen(req, timeout=8) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        return {
            "tag_name": str(payload.get("tag_name", "")).strip(),
            "name": str(payload.get("name", "")).strip(),
            "html_url": str(payload.get("html_url", "")).strip() or self.RELEASES_PAGE,
            "published_at": str(payload.get("published_at", "")).strip(),
        }

    def check_for_updates(self):
        try:
            release = self.fetch_latest_release()
            latest_tag = release["tag_name"] or release["name"] or "v0.0.0"
            latest_version = self.normalize_version(latest_tag)
            current_version = self.normalize_version(self.current_version)

            if latest_version > current_version:
                should_open = messagebox.askyesno(
                    "有可用更新",
                    (
                        f"目前版本: {self.current_version}\n"
                        f"最新版本: {latest_tag}\n\n"
                        "要前往 GitHub Release 頁面下載更新嗎？"
                    ),
                )
                self.update_log(f"檢查更新完成：發現新版本 {latest_tag}")
                if should_open:
                    webbrowser.open(release["html_url"] or self.RELEASES_PAGE)
            else:
                messagebox.showinfo("檢查更新", f"你目前使用的是最新版本 ({self.current_version})。")
                self.update_log("檢查更新完成：目前已是最新版本。")
        except url_error.URLError as exc:
            messagebox.showwarning("檢查更新失敗", f"無法連線到 GitHub：{exc}")
            self.update_log(f"檢查更新失敗（網路連線問題）：{exc}")
        except Exception as exc:
            messagebox.showwarning("檢查更新失敗", f"發生未預期錯誤：{exc}")
            self.update_log(f"檢查更新失敗：{exc}")

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as file:
                    return json.load(file)
            except Exception:
                pass
        return {"last_port": "", "last_employee_id": "", "product_history": [], "test_access_code": "2468"}

    def save_config(self):
        self.config["last_port"] = self.port_menu.get()
        self.config["last_employee_id"] = self.emp_entry.get().strip()

        current_prod = self.prod_combo.get().strip()
        if current_prod and current_prod not in self.config["product_history"]:
            self.config["product_history"].append(current_prod)
            self.config["product_history"] = self.config["product_history"][-20:]
            self.prod_combo.configure(values=self.config["product_history"])

        with open(CONFIG_FILE, "w", encoding="utf-8") as file:
            json.dump(self.config, file, indent=4, ensure_ascii=False)

    def get_ports(self):
        ports = [port.device for port in serial.tools.list_ports.comports()]
        if self.test_mode_unlocked and self.TEST_MODE_PORT not in ports:
            ports.append(self.TEST_MODE_PORT)
        return ports if ports else [self.NO_PORT_TEXT]

    def refresh_port_menu(self):
        ports = self.get_ports()
        current = self.port_menu.get()
        self.port_menu.configure(values=ports)
        if current in ports:
            self.port_menu.set(current)
        else:
            self.port_menu.set(ports[0])

    def handle_test_mode_shortcut(self, _event=None):
        self.prompt_unlock_test_mode()

    def prompt_unlock_test_mode(self):
        if self.test_mode_unlocked:
            self.update_log("測試模式已解鎖。")
            return
        dialog = ctk.CTkInputDialog(text="請輸入測試模式存取碼：", title="解鎖測試模式")
        code = dialog.get_input()
        if code is None:
            return
        expected = str(self.config.get("test_access_code", "2468"))
        if code.strip() == expected:
            self.test_mode_unlocked = True
            self.refresh_port_menu()
            self.update_log("測試模式已解鎖。")
        else:
            messagebox.showerror("錯誤", "存取碼不正確")

    def show_test_panel(self):
        self.test_weight_entry.delete(0, "end")
        self.test_weight_entry.insert(0, "10.0000")
        self.test_unit_entry.delete(0, "end")
        self.test_unit_entry.insert(0, "g")

        self.test_weight_label.pack(pady=(8, 0))
        self.test_weight_entry.pack(pady=4, padx=20)
        self.test_unit_label.pack(pady=(6, 0))
        self.test_unit_entry.pack(pady=4, padx=20)
        self.test_send_btn.pack(pady=(10, 10), padx=20)
        self.test_weight_entry.bind("<Return>", lambda _e: self.submit_test_data())
        self.test_unit_entry.bind("<Return>", lambda _e: self.submit_test_data())
        self.test_frame.pack(pady=(0, 12), padx=20, fill="x")

    def hide_test_panel(self):
        self.test_weight_entry.unbind("<Return>")
        self.test_unit_entry.unbind("<Return>")
        self.test_frame.pack_forget()

    def get_sheet_header(self, worksheet):
        if worksheet.max_row < 1:
            return []
        return ["" if cell.value is None else str(cell.value).strip() for cell in worksheet[1]]

    def is_detail_sheet(self, worksheet):
        header = set(self.get_sheet_header(worksheet))
        markers = {"員工編號", "日期", "時間", "產品編號", "名稱", "類別", "重量", "瓶", "餘量", "ID"}
        return worksheet.max_column >= 9 and len(header.intersection(markers)) >= 4

    def is_inventory_sheet(self, worksheet):
        header = self.get_sheet_header(worksheet)
        return worksheet.max_column >= 5 and len(header) >= 1 and header[0] == "ID"

    def find_detail_sheet(self, workbook):
        ws = self.find_sheet(workbook, self.DETAIL_SHEET, self.LEGACY_DETAIL_SHEET)
        if ws:
            return ws
        for sheet_name in workbook.sheetnames:
            candidate = workbook[sheet_name]
            if self.is_detail_sheet(candidate):
                return candidate
        return None

    def find_inventory_sheet(self, workbook):
        ws = self.find_sheet(workbook, self.INVENTORY_SHEET, self.LEGACY_INVENTORY_SHEET)
        if ws:
            return ws
        for sheet_name in workbook.sheetnames:
            candidate = workbook[sheet_name]
            if self.is_inventory_sheet(candidate):
                return candidate
        return None

    def find_sheet(self, workbook, primary_name, legacy_name=None):
        if primary_name in workbook.sheetnames:
            return workbook[primary_name]
        if legacy_name and legacy_name in workbook.sheetnames:
            return workbook[legacy_name]
        return None

    def rename_legacy_sheet(self, workbook, primary_name, legacy_name):
        if primary_name not in workbook.sheetnames and legacy_name in workbook.sheetnames:
            workbook[legacy_name].title = primary_name
            return True
        return False

    def ensure_headers(self, worksheet, headers):
        first_row = [cell.value for cell in worksheet[1]] if worksheet.max_row >= 1 else []
        first_row = ["" if value is None else str(value).strip() for value in first_row]

        if first_row[: len(headers)] == headers and len(first_row) == len(headers):
            return False

        data_rows = list(worksheet.iter_rows(min_row=2, values_only=True)) if worksheet.max_row >= 2 else []
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(headers)
        for row in data_rows:
            worksheet.append(list(row))
        return True

    def normalize_detail_row(self, row):
        if row is None:
            return None
        values = list(row)
        if not values or all(value is None or str(value).strip() == "" for value in values):
            return None

        # 新格式: 員工編號, 日期, 時間, 產品編號, 名稱, 類別, 重量, 瓶, 餘量
        if len(values) >= 9 and not self.looks_like_date(values[0]):
            return (
                self.safe_text(values[0]),
                self.safe_text(values[1]),
                self.safe_text(values[2]),
                self.safe_text(values[3]),
                self.safe_text(values[4]),
                self.safe_text(values[5]),
                self.safe_number(values[6]),
                self.safe_number(values[7], default=1, as_int=True),
                self.safe_number(values[8]),
            )

        # 舊格式: 日期, 時間, ID, 名稱, 類別, 重量, 單位, 瓶數, 標準, 餘量
        if len(values) >= 10:
            return (
                "",
                self.safe_text(values[0]),
                self.safe_text(values[1]),
                self.safe_text(values[2]),
                self.safe_text(values[3]),
                self.safe_text(values[4]),
                self.safe_number(values[5]),
                self.safe_number(values[7], default=1, as_int=True),
                self.safe_number(values[9]),
            )

        # 退化處理（避免舊資料長度不一致時崩潰）
        padded = values + [""] * (10 - len(values))
        return (
            "",
            self.safe_text(padded[0]),
            self.safe_text(padded[1]),
            self.safe_text(padded[2]),
            self.safe_text(padded[3]),
            self.safe_text(padded[4]),
            self.safe_number(padded[5]),
            self.safe_number(padded[7], default=1, as_int=True),
            self.safe_number(padded[9]),
        )

    def looks_like_date(self, value):
        if value is None:
            return False
        text = str(value).strip()
        return len(text) >= 8 and "-" in text and text[:4].isdigit()

    def safe_text(self, value):
        return "" if value is None else str(value).strip()

    def safe_number(self, value, default=0, as_int=False):
        try:
            num = float(value)
        except Exception:
            num = float(default)
        if as_int:
            return int(round(num))
        return round(num, 4)

    def upgrade_workbook_structure(self):
        if not os.path.exists(self.file_name):
            return

        changed = False
        try:
            wb = load_workbook(self.file_name)
            changed |= self.rename_legacy_sheet(wb, self.DETAIL_SHEET, self.LEGACY_DETAIL_SHEET)
            changed |= self.rename_legacy_sheet(wb, self.INVENTORY_SHEET, self.LEGACY_INVENTORY_SHEET)

            ws_detail = self.find_detail_sheet(wb)
            if ws_detail and ws_detail.title != self.DETAIL_SHEET and self.DETAIL_SHEET not in wb.sheetnames:
                ws_detail.title = self.DETAIL_SHEET
                changed = True
            if ws_detail is None:
                ws_detail = wb.create_sheet(self.DETAIL_SHEET)
                ws_detail.append(self.DETAIL_HEADERS)
                changed = True

            ws_inv = self.find_inventory_sheet(wb)
            if ws_inv and ws_inv.title != self.INVENTORY_SHEET and self.INVENTORY_SHEET not in wb.sheetnames:
                ws_inv.title = self.INVENTORY_SHEET
                changed = True
            if ws_inv is None:
                ws_inv = wb.create_sheet(self.INVENTORY_SHEET)
                ws_inv.append(self.INVENTORY_HEADERS)
                changed = True

            changed |= self.migrate_detail_sheet(ws_detail)
            changed |= self.ensure_headers(ws_inv, self.INVENTORY_HEADERS)

            if changed:
                wb.save(self.file_name)
                self.update_log("已自動升級 Excel 結構：詳細記錄已加入員工編號欄位。")
            wb.close()
        except Exception as exc:
            self.update_log(f"自動升級 Excel 結構失敗: {exc}")

    def migrate_detail_sheet(self, worksheet):
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            worksheet.append(self.DETAIL_HEADERS)
            return True

        header = ["" if value is None else str(value).strip() for value in rows[0]]
        already_new = header[: len(self.DETAIL_HEADERS)] == self.DETAIL_HEADERS and len(header) == len(
            self.DETAIL_HEADERS
        )
        if already_new:
            return False

        converted = []
        for row in rows[1:]:
            normalized = self.normalize_detail_row(row)
            if normalized:
                converted.append(list(normalized))

        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(self.DETAIL_HEADERS)
        for row in converted:
            worksheet.append(row)
        return True

    def get_standard_from_excel(self, cid):
        if not os.path.exists(self.file_name):
            return None
        try:
            wb = load_workbook(self.file_name, read_only=True)
            ws = self.find_inventory_sheet(wb)
            if ws:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == str(cid):
                        wb.close()
                        return float(row[3]) if row[3] is not None else 0.0
            wb.close()
        except Exception:
            pass
        return None

    def toggle_monitoring(self):
        if not self.is_monitoring:
            self.start_monitoring()
        else:
            self.stop_monitoring()

    def start_monitoring(self):
        port = self.port_menu.get()
        if port == self.NO_PORT_TEXT:
            self.update_log("錯誤: 請先選擇有效的 COM Port")
            return
        if port == self.TEST_MODE_PORT:
            if not self.test_mode_unlocked:
                messagebox.showwarning("警告", "你沒有測試模式權限。")
                return
            self.start_test_mode()
            return
        try:
            self.ser = serial.Serial(port, 9600, timeout=1)
            self.is_monitoring = True
            self.is_test_mode = False
            self.save_config()
            self.btn_start.configure(text="停止監聽", fg_color="red")
            self.update_log(f"系統已開始監聽 ({port})")
            threading.Thread(target=self.listen_serial, daemon=True).start()
        except Exception as exc:
            self.update_log(f"啟動監聽失敗: {exc}")

    def start_test_mode(self):
        self.is_monitoring = True
        self.is_test_mode = True
        self.save_config()
        self.btn_start.configure(text="停止監聽", fg_color="red")
        self.show_test_panel()
        self.update_log("已進入測試模式，請在左側輸入測試重量並按下「送出測試數據」。")

    def stop_monitoring(self):
        self.is_monitoring = False
        self.is_test_mode = False
        if self.ser:
            self.ser.close()
            self.ser = None
        self.hide_test_panel()
        self.btn_start.configure(text="開始監聽", fg_color="green")
        self.update_log("監聽已停止")

    def listen_serial(self):
        while self.is_monitoring:
            try:
                if self.ser and self.ser.in_waiting > 0:
                    raw_line = self.ser.readline().decode("ascii", errors="replace").strip()
                    if raw_line:
                        self.after(0, self.handle_weight_data, raw_line)
            except Exception:
                break

    def submit_test_data(self):
        if not (self.is_monitoring and self.is_test_mode):
            messagebox.showwarning("警告", "請先進入測試模式。")
            return
        weight_text = self.test_weight_entry.get().strip()
        unit_text = self.test_unit_entry.get().strip() or "g"
        try:
            weight_value = float(weight_text)
        except Exception:
            messagebox.showerror("錯誤", "測試重量格式錯誤，請輸入數字。")
            return

        raw_line = f"S S {weight_value:.4f} {unit_text}"
        self.handle_weight_data(raw_line)

    def handle_weight_data(self, raw_line):
        parts = raw_line.split()
        if not (len(parts) >= 4 and parts[0] in ["S", "SD"]):
            if raw_line:
                self.update_log(f"略過無效秤重資料: {raw_line}")
            return

        weight = parts[2]
        unit = parts[3]

        self.weight_display.configure(text=weight)
        self.unit_label.configure(text=unit)

        employee_id = self.emp_entry.get().strip()
        cid = self.id_entry.get().strip() or "0000"
        prod = self.prod_combo.get().strip() or "未命名"
        ctype = self.type_var.get()
        bot = self.bot_entry.get().strip()

        std_val = self.get_standard_from_excel(cid)
        if std_val is None:
            self.popup_standard_input(employee_id, cid, prod, ctype, weight, unit, bot, default_std=weight)
        else:
            self.process_save(employee_id, cid, prod, ctype, weight, unit, bot, std_val)

    def update_log(self, message):
        self.log_box.insert("end", message + "\n" + "-" * 30 + "\n")
        self.log_box.see("end")

    def popup_standard_input(self, employee_id, cid, prod, ctype, weight, unit, bot, default_std):
        dialog = ctk.CTkToplevel(self)
        dialog.title("首次輸入標準")
        dialog.geometry("360x220")
        dialog.attributes("-topmost", True)
        dialog.grab_set()

        label = ctk.CTkLabel(
            dialog,
            text=f"產品編號 [{cid}] 首次出現\n請輸入安全庫存標準:",
            font=ctk.CTkFont(size=16),
        )
        label.pack(pady=(20, 10))

        entry = ctk.CTkEntry(dialog, width=150, font=ctk.CTkFont(size=16))
        entry.pack(pady=5)
        entry.insert(0, str(default_std))

        def confirm():
            value = entry.get()
            dialog.destroy()
            try:
                float_val = float(value)
            except Exception:
                float_val = float(default_std)
            self.process_save(employee_id, cid, prod, ctype, weight, unit, bot, float_val)

        button = ctk.CTkButton(dialog, text="確認儲存", command=confirm)
        button.pack(pady=15)

    def modify_standard(self):
        cid = self.id_entry.get().strip()
        std_val = self.get_standard_from_excel(cid)

        if std_val is None:
            messagebox.showwarning("警告", f"產品編號 [{cid}] 尚未建立紀錄，請先秤重一次。")
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("修改安全庫存標準")
        dialog.geometry("360x220")
        dialog.attributes("-topmost", True)
        dialog.grab_set()

        label = ctk.CTkLabel(
            dialog,
            text=f"目前產品編號 [{cid}] 的標準為: {std_val}\n請輸入新標準:",
            font=ctk.CTkFont(size=14),
        )
        label.pack(pady=(20, 10))

        entry = ctk.CTkEntry(dialog, width=150)
        entry.pack(pady=5)
        entry.insert(0, str(std_val))

        def confirm():
            value = entry.get()
            dialog.destroy()
            try:
                new_std = float(value)
                self.update_standard_only(cid, new_std)
            except Exception:
                messagebox.showerror("錯誤", "請輸入有效數字")

        button = ctk.CTkButton(dialog, text="確認修改", command=confirm)
        button.pack(pady=15)

    def update_standard_only(self, cid, new_std):
        try:
            wb = load_workbook(self.file_name)
            ws_inv = self.find_inventory_sheet(wb)
            if ws_inv is None:
                self.update_log("修改失敗: 找不到庫存餘量工作表。")
                wb.close()
                return

            alert_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            normal_fill = PatternFill(fill_type=None)

            for row in range(2, ws_inv.max_row + 1):
                if str(ws_inv.cell(row=row, column=1).value) == str(cid):
                    ws_inv.cell(row=row, column=4, value=new_std)
                    try:
                        current_rem = float(ws_inv.cell(row=row, column=5).value)
                    except Exception:
                        current_rem = 0.0

                    fill = alert_fill if current_rem < new_std else normal_fill
                    for col in range(1, 6):
                        ws_inv.cell(row=row, column=col).fill = fill
                    break
            wb.save(self.file_name)
            wb.close()
            self.update_log(f"--- 產品編號 [{cid}] 標準已更新為 {new_std} ---")
        except PermissionError:
            self.update_log("!!! 修改失敗: 請先關閉 Excel 檔案 !!!")

    def process_save(self, employee_id, cid, prod, ctype, weight, unit, bot, std_val):
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        employee_display = employee_id if employee_id else "未填"

        log_msg = (
            f"[{time_str}] 員工:{employee_display} | {ctype} | "
            f"{prod}({cid}) x{bot}瓶: {weight} {unit} (標準:{std_val})"
        )
        self.update_log(log_msg)

        self.save_to_excel(employee_display, cid, prod, ctype, date_str, time_str, weight, unit, bot, std_val)
        self.save_config()
        self.load_recent_records()

    def save_to_excel(self, employee_id, cid, prod, ctype, d, t, w, u, bot_str, std_val):
        try:
            try:
                w_val = float(w)
            except Exception:
                w_val = 0.0

            try:
                bottles = int(bot_str)
            except Exception:
                bottles = 1
            if bottles <= 0:
                bottles = 1

            alert_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            normal_fill = PatternFill(fill_type=None)

            if not os.path.exists(self.file_name):
                wb = Workbook()
                ws_log = wb.active
                ws_log.title = self.DETAIL_SHEET
                ws_log.append(self.DETAIL_HEADERS)

                ws_inv = wb.create_sheet(title=self.INVENTORY_SHEET)
                ws_inv.append(self.INVENTORY_HEADERS)
                wb.save(self.file_name)
                wb.close()

            self.upgrade_workbook_structure()
            wb = load_workbook(self.file_name)

            ws_log = self.find_detail_sheet(wb)
            if ws_log is None:
                ws_log = wb.create_sheet(self.DETAIL_SHEET)
                ws_log.append(self.DETAIL_HEADERS)

            ws_inv = self.find_inventory_sheet(wb)
            if ws_inv is None:
                ws_inv = wb.create_sheet(self.INVENTORY_SHEET)
                ws_inv.append(self.INVENTORY_HEADERS)

            match_row = None
            current_rem = 0.0

            for row in range(2, ws_inv.max_row + 1):
                if str(ws_inv.cell(row=row, column=1).value) == str(cid):
                    match_row = row
                    val = ws_inv.cell(row=row, column=5).value
                    if val is not None:
                        try:
                            current_rem = float(val)
                        except Exception:
                            current_rem = 0.0
                    break

            total_change = w_val * bottles
            if ctype == self.OUT_TYPE:
                total_change = -total_change
            new_rem = current_rem + total_change

            if match_row:
                ws_inv.cell(row=match_row, column=2, value=prod)
                ws_inv.cell(row=match_row, column=3, value=u)
                ws_inv.cell(row=match_row, column=4, value=std_val)
                ws_inv.cell(row=match_row, column=5, value=new_rem)
            else:
                ws_inv.append([cid, prod, u, std_val, new_rem])
                match_row = ws_inv.max_row

            fill = alert_fill if new_rem < std_val else normal_fill
            for col in range(1, 6):
                ws_inv.cell(row=match_row, column=col).fill = fill

            ws_log.append([employee_id, d, t, cid, prod, ctype, w_val, bottles, round(new_rem, 4)])
            wb.save(self.file_name)
            wb.close()
        except PermissionError:
            self.update_log("!!! 儲存失敗: 請先關閉 Excel 檔案 !!!")
        except Exception as exc:
            self.update_log(f"儲存資料時發生錯誤: {exc}")


if __name__ == "__main__":
    app = BalanceGUI()
    app.mainloop()
